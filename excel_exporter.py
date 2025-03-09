import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.drawing.image import Image
import io
from PIL import Image as PILImage
from PIL import ImageDraw, ImageFont
import logging

logger = logging.getLogger(__name__)

def generate_cover_image(title, subtitle, output_dir):
    """Generate a stylish cover image for the Excel report."""
    try:
        # Create image with gradient background
        width, height = 800, 400
        image = PILImage.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(image)
        
        # Draw gradient background (light blue to white)
        for y in range(height):
            # Create gradient from top to bottom
            r = int(237 + (255 - 237) * y / height)
            g = int(242 + (255 - 242) * y / height)
            b = int(247 + (255 - 247) * y / height)
            draw.line([(0, y), (width, y)], fill=(r, g, b))
        
        # Try to use a nice font, fall back to default if not available
        try:
            title_font = ImageFont.truetype("Arial Bold", 40)
            subtitle_font = ImageFont.truetype("Arial", 24)
        except IOError:
            # Use default font if custom font not available
            title_font = ImageFont.load_default()
            subtitle_font = ImageFont.load_default()
        
        # Draw title and subtitle
        draw.text((width//2, height//2 - 40), title, fill=(41, 65, 148), font=title_font, anchor="mm")
        draw.text((width//2, height//2 + 20), subtitle, fill=(100, 116, 139), font=subtitle_font, anchor="mm")
        
        # Draw accent line
        draw.rectangle([(width//2 - 100, height//2 - 5), (width//2 + 100, height//2 - 3)], fill=(41, 65, 148))
        
        # Add current date
        date_font = ImageFont.truetype("Arial", 16) if 'Arial' in ImageFont.truetype.__self__.keys() else ImageFont.load_default()
        current_date = datetime.now().strftime("%B %d, %Y")
        draw.text((width//2, height - 40), f"Generated on {current_date}", fill=(100, 116, 139), font=date_font, anchor="mm")
        
        # Save image
        image_path = os.path.join(output_dir, 'report_cover.png')
        image.save(image_path)
        return image_path
    except Exception as e:
        logger.error(f"Error generating cover image: {str(e)}")
        return None

def generate_excel_report(documents, output_dir):
    """Generate a visually enhanced Excel report from extracted document text and paragraphs."""
    try:
        logger.info("Generating enhanced Excel report")
        
        # Create a new workbook with openpyxl
        wb = Workbook()
        
        # Create cover sheet
        cover_sheet = wb.active
        cover_sheet.title = "Cover"
        
        # Try to add cover image
        cover_image_path = generate_cover_image("Document Analysis Report", "Content and Paragraph Analysis", output_dir)
        if cover_image_path and os.path.exists(cover_image_path):
            img = Image(cover_image_path)
            img.width, img.height = 600, 300
            cover_sheet.add_image(img, 'B2')
            # Delete the temporary image file
            os.remove(cover_image_path)
        
        # Add report information to cover page
        cover_sheet['B15'] = "Report Information"
        cover_sheet['B15'].font = Font(size=16, bold=True, color="2C3E50")
        
        info_rows = [
            ["Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Documents Analyzed:", len(documents)],
            ["Total Pages:", sum(doc.page_count for doc in documents if doc.page_count)],
            ["Total Paragraphs:", sum(doc.paragraph_count for doc in documents if doc.paragraph_count)]
        ]
        
        for idx, (label, value) in enumerate(info_rows, 16):
            cover_sheet[f'B{idx}'] = label
            cover_sheet[f'B{idx}'].font = Font(bold=True)
            cover_sheet[f'C{idx}'] = value
        
        # Define styles
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin', color="BFBFBF"),
            right=Side(style='thin', color="BFBFBF"),
            top=Side(style='thin', color="BFBFBF"),
            bottom=Side(style='thin', color="BFBFBF")
        )
        center_aligned = Alignment(horizontal='center', vertical='center')
        wrapped_alignment = Alignment(wrap_text=True, vertical='top')
        
        # Create dashboard sheet
        dashboard = wb.create_sheet("Dashboard")
        
        # Prepare data for dashboard charts
        file_types = {}
        doc_sizes = []
        paragraph_counts = []
        
        for doc in documents:
            # Count file types
            file_type = doc.file_type.upper()
            file_types[file_type] = file_types.get(file_type, 0) + 1
            
            # Collect document sizes for size distribution
            doc_sizes.append((doc.original_filename, doc.file_size / 1024))  # Size in KB
            
            # Collect paragraph counts
            paragraph_counts.append((doc.original_filename, doc.paragraph_count or 0))
        
        # Add file type distribution chart (pie chart)
        dashboard['A1'] = "File Type Distribution"
        dashboard['A1'].font = Font(size=14, bold=True)
        
        # Write file type data
        dashboard['A3'] = "File Type"
        dashboard['B3'] = "Count"
        dashboard['A3'].font = header_font
        dashboard['B3'].font = header_font
        dashboard['A3'].fill = header_fill
        dashboard['B3'].fill = header_fill
        
        row = 4
        for file_type, count in file_types.items():
            dashboard[f'A{row}'] = file_type
            dashboard[f'B{row}'] = count
            row += 1
        
        # Create pie chart
        pie = PieChart()
        labels = Reference(dashboard, min_col=1, min_row=4, max_row=row-1)
        data = Reference(dashboard, min_col=2, min_row=3, max_row=row-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Document Types"
        
        # Style the pie chart
        pie.height = 10  # height in cm
        pie.width = 15   # width in cm
        
        # Add the chart to the worksheet
        dashboard.add_chart(pie, "D3")
        
        # Add document size chart (horizontal bar chart)
        dashboard['A15'] = "Document Size Distribution"
        dashboard['A15'].font = Font(size=14, bold=True)
        
        # Write size data
        dashboard['A17'] = "Document"
        dashboard['B17'] = "Size (KB)"
        dashboard['A17'].font = header_font
        dashboard['B17'].font = header_font
        dashboard['A17'].fill = header_fill
        dashboard['B17'].fill = header_fill
        
        # Sort documents by size (largest first)
        doc_sizes.sort(key=lambda x: x[1], reverse=True)
        doc_sizes = doc_sizes[:10]  # Top 10 documents by size
        
        row = 18
        for doc_name, size in doc_sizes:
            dashboard[f'A{row}'] = doc_name
            dashboard[f'B{row}'] = size
            row += 1
        
        # Create bar chart for document sizes
        bar = BarChart()
        bar.type = "bar"  # Horizontal bar
        bar.title = "Document Sizes (Top 10)"
        bar.y_axis.title = "Document"
        bar.x_axis.title = "Size (KB)"
        
        data = Reference(dashboard, min_col=2, min_row=17, max_row=row-1)
        cats = Reference(dashboard, min_col=1, min_row=18, max_row=row-1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        
        # Style the bar chart
        bar.height = 15  # height in cm
        bar.width = 20   # width in cm
        
        # Add the chart to the worksheet
        dashboard.add_chart(bar, "D17")
        
        # Create summary sheet with enhanced styling
        summary = wb.create_sheet("Summary")
        
        # Prepare summary data
        summary_data = []
        for doc in documents:
            summary_data.append({
                'ID': doc.id,
                'Filename': doc.original_filename,
                'File Type': doc.file_type.upper(),
                'File Size (KB)': round(doc.file_size / 1024, 2),
                'Page Count': doc.page_count or 0,
                'Paragraph Count': doc.paragraph_count or 0,
                'Upload Date': doc.upload_date.strftime("%Y-%m-%d %H:%M") if doc.upload_date else "",
                'Status': doc.status.capitalize() if doc.status else "",
                'Text Length': len(doc.extracted_text) if doc.extracted_text else 0,
            })
        
        # Write summary headers
        headers = list(summary_data[0].keys()) if summary_data else []
        for col_idx, header in enumerate(headers, 1):
            cell = summary.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_aligned
        
        # Write summary data with alternating row colors
        for row_idx, data_row in enumerate(summary_data, 2):
            # Apply alternate row styling
            row_fill = alt_row_fill if row_idx % 2 == 0 else None
            
            for col_idx, key in enumerate(headers, 1):
                cell = summary.cell(row=row_idx, column=col_idx, value=data_row[key])
                cell.border = border
                if row_fill:
                    cell.fill = row_fill
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            
            for row_idx in range(2, len(summary_data) + 2):
                cell_value = summary.cell(row=row_idx, column=col_idx).value
                max_length = max(max_length, len(str(cell_value) if cell_value is not None else ""))
            
            adjusted_width = max_length + 4  # Add padding
            summary.column_dimensions[column_letter].width = min(50, adjusted_width)  # Cap width at 50
        
        # Create text extract sheet
        text_sheet = wb.create_sheet("Full Text")
        
        # Add headers
        headers = ["ID", "Filename", "Pages", "Extracted Text"]
        for col_idx, header in enumerate(headers, 1):
            cell = text_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_aligned
        
        # Set column widths
        text_sheet.column_dimensions['A'].width = 5
        text_sheet.column_dimensions['B'].width = 40
        text_sheet.column_dimensions['C'].width = 8
        text_sheet.column_dimensions['D'].width = 120
        
        # Write text data
        for row_idx, doc in enumerate(documents, 2):
            text_sheet.cell(row=row_idx, column=1, value=doc.id).border = border
            text_sheet.cell(row=row_idx, column=2, value=doc.original_filename).border = border
            text_sheet.cell(row=row_idx, column=3, value=doc.page_count).border = border
            
            # Truncate text if necessary
            text_cell = text_sheet.cell(row=row_idx, column=4)
            text_cell.value = doc.extracted_text[:32000] if doc.extracted_text else "No text extracted"
            text_cell.border = border
            text_cell.alignment = wrapped_alignment
            
            # Apply alternate row styling
            if row_idx % 2 == 0:
                for col in range(1, 5):
                    text_sheet.cell(row=row_idx, column=col).fill = alt_row_fill
            
            # Set row height based on content
            text_sheet.row_dimensions[row_idx].height = 120
        
        # Create paragraph sheet with enhanced styles
        para_sheet = wb.create_sheet("Paragraphs")
        
        # Add headers
        headers = ["ID", "Document Count", "Content", "Documents"]
        for col_idx, header in enumerate(headers, 1):
            cell = para_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_aligned
        
        # Set column widths
        para_sheet.column_dimensions['A'].width = 5
        para_sheet.column_dimensions['B'].width = 15
        para_sheet.column_dimensions['C'].width = 100
        para_sheet.column_dimensions['D'].width = 40
        
        # Get all unique paragraphs from all documents
        all_paragraphs = set()
        for doc in documents:
            for para in doc.paragraphs:
                all_paragraphs.add(para)
        
        # Sort paragraphs by number of documents (most shared first)
        sorted_paragraphs = sorted(all_paragraphs, key=lambda p: len(p.documents), reverse=True)
        
        # Write paragraph data
        for row_idx, para in enumerate(sorted_paragraphs, 2):
            # Get a list of document filenames this paragraph appears in
            doc_list = [doc.original_filename for doc in para.documents if doc in documents]
            
            para_sheet.cell(row=row_idx, column=1, value=para.id).border = border
            para_sheet.cell(row=row_idx, column=2, value=len(doc_list)).border = border
            
            # Truncate paragraph content for Excel if extremely long
            content_cell = para_sheet.cell(row=row_idx, column=3)
            content = para.content
            if len(content) > 32000:  # Excel cell limit
                content = content[:32000] + "... (truncated)"
            content_cell.value = content
            content_cell.border = border
            content_cell.alignment = wrapped_alignment
            
            doc_cell = para_sheet.cell(row=row_idx, column=4)
            doc_cell.value = ", ".join(doc_list)
            doc_cell.border = border
            doc_cell.alignment = wrapped_alignment
            
            # Apply alternate row styling and highlight shared paragraphs
            if len(doc_list) > 1:
                # Shared paragraph - light highlight
                highlight_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                for col in range(1, 5):
                    para_sheet.cell(row=row_idx, column=col).fill = highlight_fill
            elif row_idx % 2 == 0:
                # Alternate row styling for non-shared paragraphs
                for col in range(1, 5):
                    para_sheet.cell(row=row_idx, column=col).fill = alt_row_fill
            
            # Set row height based on content
            para_sheet.row_dimensions[row_idx].height = 100
        
        # Freeze the header row in all sheets
        summary.freeze_panes = summary["A2"]
        text_sheet.freeze_panes = text_sheet["A2"]
        para_sheet.freeze_panes = para_sheet["A2"]
        
        # Generate timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f'document_analysis_{timestamp}.xlsx')
        
        # Save the workbook
        wb.save(filename)
        
        logger.info(f"Enhanced Excel report generated successfully: {filename}")
        return os.path.basename(filename)
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        raise